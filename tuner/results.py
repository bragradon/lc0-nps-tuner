import csv
import shutil
from pathlib import Path
from re import Match

from openpyxl import Workbook, load_workbook

from tuner.environment import Environment

RECORDERS = {}


def get_results_recorder(environment: Environment) -> "ResultsRecorder":
    return RECORDERS[environment.results_file_format](environment)


def register(filetype):
    def inner(recorder):
        RECORDERS[filetype] = recorder

    return inner


class ResultsRecorder:
    def __init__(self, environment: Environment):
        self.environment = environment

    def record_result(self, *args, **kwargs):
        raise NotImplementedError

    @staticmethod
    def extract_results(results: Match):
        if results is not None:
            depth, seldepth, nps = (
                results.group("depth"),
                results.group("seldepth"),
                results.group("nps"),
            )
        else:
            depth, seldepth, nps = None, None, None

        return depth, seldepth, nps

    def complete_run(self, config_file):
        shutil.move(str(config_file.absolute()), config_file.parent / "processed")

    def read_run_config(self, config_file):
        with open(config_file) as config:
            config = config.read().replace('--', '').strip().split('\n')
            return dict(sorted(i.split('=', 1) for i in config))


@register("xlsx")
class XLSXResultsRecorder(ResultsRecorder):
    def __init__(self, environment: Environment):
        super().__init__(environment)

        self.wb, self.ws = self.get_results_excel()
        self.add_results_workbook_headers(self.ws)
        self.save_workbook()

    def get_results_excel(self):
        results = self.environment.results_path.absolute()

        try:
            wb = load_workbook(str(results))
        except FileNotFoundError:
            wb = Workbook()

        if "Results" not in wb.sheetnames:
            ws = wb.create_sheet(title="Results", index=0)
        else:
            ws = wb["Results"]

        return wb, ws

    def add_results_workbook_headers(self, worksheet):
        worksheet["A1"] = "Best NPS"
        worksheet["B1"] = "Filename"
        worksheet["A2"] = "=MAX(E:E)"
        worksheet["B2"] = '=INDIRECT("D" & MATCH(MAX(E:E), E:E, 0))'

        worksheet["A4"] = "Best Depth"
        worksheet["B4"] = "Filename"
        worksheet["A5"] = "=MAX(F:F)"
        worksheet["B5"] = '=INDIRECT("D" & MATCH(MAX(F:F), F:F, 0))'

        worksheet["A7"] = "Best SelDepth"
        worksheet["B7"] = "Filename"
        worksheet["A8"] = "=MAX(G:G)"
        worksheet["B8"] = '=INDIRECT("D" & MATCH(MAX(G:G), G:G, 0))'

        worksheet["D1"] = "LC0 Parameters Run Results"
        worksheet["D2"] = "Filename"
        worksheet["E2"] = "NPS"
        worksheet["F2"] = "DEPTH"
        worksheet["G2"] = "SELDEPTH"

        for i, header in enumerate(sorted(self.environment.options['options'].keys())):
            cellref = worksheet.cell(row=2, column=i + ord("h") - ord("a") + 1)
            cellref.value = header

    def save_workbook(self):
        self.wb.save(filename=str(self.environment.results_path.absolute()))

    def record_result(self, results: Match, config_file: Path):
        depth, seldepth, nps = self.extract_results(results)
        config_filename = str(config_file.name)
        config = self.read_run_config(config_file)

        row = int(config_filename.split(".")[0])
        self.ws.cell(row + 2, ord("d") - ord("a") + 1, config_filename)
        self.ws.cell(row + 2, ord("e") - ord("a") + 1, int(nps if nps else 0))
        self.ws.cell(row + 2, ord("f") - ord("a") + 1, int(depth if depth else 0))
        self.ws.cell(row + 2, ord("g") - ord("a") + 1, int(seldepth if seldepth else 0))

        for i, (name, value) in enumerate(config.items()):
            self.ws.cell(row + 2, ord("h") - ord("a") + 1 + i, value)

        self.save_workbook()
        self.complete_run(config_file)


@register("csv")
class CSVResultsRecorder(ResultsRecorder):
    def __init__(self, environment: Environment):
        super().__init__(environment)

        file_path = str(self.environment.results_path.absolute())

        header = ["Filename", "NPS", "DEPTH", "SELDEPTH"]
        header += sorted(self.environment.options['options'].keys())

        self.file = open(file_path, "w")
        self.writer = csv.DictWriter(
            self.file, fieldnames=header
        )
        self.writer.writeheader()

    def __del__(self):
        self.file.close()

    def record_result(self, results: Match, config_file: Path):
        depth, seldepth, nps = self.extract_results(results)
        config = self.read_run_config(config_file)
        
        row = {
            "Filename": str(config_file.name),
            "NPS": nps,
            "DEPTH": depth,
            "SELDEPTH": seldepth,
            **config,
        }
        self.writer.writerow(row)
        self.complete_run(config_file)
